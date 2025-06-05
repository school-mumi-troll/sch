import nest_asyncio
import openpyxl
import asyncio
import logging
import re
from datetime import datetime

# Enable logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logging.getLogger("httpx").setLevel(logging.WARNING)
logger = logging.getLogger(__name__)
nest_asyncio.apply()

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    ContextTypes,
    CallbackQueryHandler,
    ConversationHandler,
)

# Состояния разговора
START, CHOOSE_CLASS, CHOOSE_STUDENT, CHOOSE_DAY = range(4)
schedule_10 = {}
schedule_11 = {}
schedule_789 = {}


def load_schedule(file_path, class_type):
    try:
        book = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sh = book.active

        # Определение дней недели и диапазонов строк
        
        week_ranges = {
            "понедельник": range(3, 12),
            "вторник": range(13, 22),
            "среда": range(23, 32),
            "четверг": range(33, 42),
            "пятница": range(43, 53)
        }
        if class_type == "789":
            week_ranges = {
                "понедельник": range(3, 12),
                "вторник": range(14, 22),
                "среда": range(24, 32),
                "четверг": range(34, 42),
                "пятница": range(45, 53)
            }


        # Собираем имена учеников
        students = []
        col = 3
        while True:
            cell_value = sh.cell(row=2, column=col).value
            if cell_value is None or cell_value == "":
                break
            students.append(cell_value)
            col += 1

        # Собираем расписание
        schedule = {}
        for student in students:
            student_schedule = {}
            col_index = students.index(student) + 3

            for day, rows in week_ranges.items():
                lessons = []
                for row in rows:
                    cell_value = sh.cell(row=row, column=col_index).value
                    lessons.append(cell_value if cell_value not in (None, "") else "-")
                student_schedule[day] = lessons
            schedule[str(student)] = student_schedule

        return schedule

    except Exception as e:
        logger.error(f"Ошибка загрузки файла {class_type}: {str(e)}")
        return {}


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    keyboard = [
        [InlineKeyboardButton("10 класс", callback_data="class_10")],
        [InlineKeyboardButton("11 класс", callback_data="class_11")],
        [InlineKeyboardButton("7-9 классы", callback_data="class_789")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    if update.message:
        await update.message.reply_text(
            "Привет! Выбери класс:",
            reply_markup=reply_markup
        )
    else:
        await update.callback_query.edit_message_text(
            "Привет! Выбери класс:",
            reply_markup=reply_markup
        )
    return CHOOSE_CLASS


async def choose_class(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()
    class_choice = query.data

    if query.data == "back_to_start":
        return await start(update, context)
    # if class_choice == "class_7_9":
    #     await query.edit_message_text("Функционал для 7-9 классов в разработке 🛠")
    #     return await start(update, context)

    context.user_data['class'] = class_choice.split("_")[1]

    # Выбираем нужное расписание
    if context.user_data['class'] == "10":
        schedule = schedule_10
    elif context.user_data['class'] == "11":
        schedule = schedule_11
    elif context.user_data['class'] == "789":
        schedule = schedule_789


    if not schedule:
        await query.edit_message_text(
            f"Расписание для {context.user_data['class']} класса временно недоступно 😢"
        )
        return await start(update, context)

    # Создаем кнопки для учеников
    keyboard = []
    students = list(schedule.keys())
    for student in students:
        keyboard.append([InlineKeyboardButton(student, callback_data=f"student_{student}")])

    keyboard.append([InlineKeyboardButton("◀ Назад", callback_data="back_to_start")])
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        f"Выбран {context.user_data['class']} класс. Выбери ученика:",
        reply_markup=reply_markup
    )
    return CHOOSE_STUDENT


async def choose_student(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    if query.data == "back_to_start":
        return await start(update, context)

    student = query.data.split("_", 1)[1]
    context.user_data['student'] = student

    # Определяем текущий день недели
    days = ["понедельник", "вторник", "среда", "четверг", "пятница"]
    today = days[datetime.today().weekday()] if datetime.today().weekday() < 5 else "понедельник"

    # Создаем кнопки для дней недели
    keyboard = []
    for day in days:
        prefix = "✅ " if day == today else ""
        keyboard.append([InlineKeyboardButton(f"{prefix}{day.capitalize()}", callback_data=f"day_{day}")])

    keyboard.append([
        # InlineKeyboardButton("◀ Назад", callback_data=f"back_to_class_{context.user_data['class']}"),
        # InlineKeyboardButton("◀ Назад", callback_data="back_to_start"),
        # InlineKeyboardButton("🏠 В начало", callback_data="back_to_start")
        InlineKeyboardButton("🏠 В начало", callback_data="back_to_start")
    ])
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        f"Ученик: {student}\nВыбери день недели:",
        reply_markup=reply_markup
    )
    return CHOOSE_DAY


async def show_schedule(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    if query.data == "back_to_start":
        return await start(update, context)
    if query.data == "back_to_days":
        return await choose_student(update, context)

    if query.data.startswith("back_to_class"):
        class_type = query.data.split("_")[-1]
        context.user_data['class'] = class_type
        return await choose_class(update, context)
    day = query.data.split("_", 1)[1]

    class_type = context.user_data['class']
    student = context.user_data['student']

    # Выбираем нужное расписание
    if context.user_data['class'] == "10":
        schedule = schedule_10
    elif context.user_data['class'] == "11":
        schedule = schedule_11
    elif context.user_data['class'] == "789":
        schedule = schedule_789
    lessons = schedule.get(student, {}).get(day, [])

    # Форматируем расписание
    schedule_text = []
    for i, lesson in enumerate(lessons, 1):
        if lesson and lesson != "-":
            # Очищаем от лишних символов
            clean_lesson = re.sub(r'[^\w\sа-яА-ЯёЁ.,!?;-]', '', str(lesson))
            schedule_text.append(f"{i}. {clean_lesson}")

    if not schedule_text:
        schedule_text = ["Занятий нет 🎉"]

    # Создаем кнопки для навигации
    keyboard = [
        [
            # InlineKeyboardButton("🔁 Выбрать другой день", callback_data="back_to_days")
            # InlineKeyboardButton("🔁 Выбрать другой день", callback_data="back_to_start")
        ],
        [
            # InlineKeyboardButton("◀ Выбрать другого ученика",
            #                      callback_data=f"back_to_class_{class_type}"),

            # InlineKeyboardButton("◀ Выбрать другого ученика",
            #                      callback_data=f"back_to_start"),
            InlineKeyboardButton("🏠 В начало", callback_data="back_to_start")
        ]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)

    await query.edit_message_text(
        f"📅 Расписание для {student} ({day}):\n\n" + "\n".join(schedule_text),
        reply_markup=reply_markup
    )
    return CHOOSE_DAY


async def handle_back(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    if query.data == "back_to_start":
        return await start(update, context)

    if query.data == "back_to_days":
        return await choose_student(update, context)

    if query.data.startswith("back_to_class"):
        class_type = query.data.split("_")[-1]
        context.user_data['class'] = class_type
        return await choose_class(update, context)

    return START


async def error_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    logger.error(f"Ошибка: {context.error}")
    if update.callback_query:
        await update.callback_query.message.reply_text("Произошла ошибка. Попробуйте снова.")
    await start(update, context)


def main() -> None:
    global schedule_10, schedule_11, schedule_789

    # Загрузка расписаний
    schedule_10 = load_schedule(r"10.xlsx", "10")
    schedule_11 = load_schedule(r"11.xlsx", "11")
    schedule_789 = load_schedule(r"789.xlsx", "789")

    # Создаем приложение
    application = ApplicationBuilder().token("7572980109:AAHLudwc2xJ8in0Kv9cpv6gb8DLdCUOHtCs").build()

    # Обработчики диалога
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            CHOOSE_CLASS: [CallbackQueryHandler(choose_class)],
            CHOOSE_STUDENT: [CallbackQueryHandler(choose_student)],
            CHOOSE_DAY: [CallbackQueryHandler(show_schedule)]
        },
        fallbacks=[
            CallbackQueryHandler(handle_back, pattern="^back_to"),
            CommandHandler("start", start)
        ],
        allow_reentry=True
    )

    application.add_handler(conv_handler)
    application.add_error_handler(error_handler)

    # Запуск бота
    application.run_polling()


if __name__ == '__main__':
    main()
